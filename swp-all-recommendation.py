import requests
import csv
import argparse
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = 'https://app.deepsecurity.trendmicro.com/api'
API_KEY  = '8BEC6851-96F0-D7D3-5002-C1062BC9E89C:CB77AF3B-16F2-F519-1E61-AFF9B2B2CF46:qKhC1TxhgCVVze6qISiL2oJq1f9GcEz+KJH2iBVUVBE='

headers = {
    'api-secret-key': API_KEY,
    'api-version': 'v1',
    'Content-Type': 'application/json'
}

MODULE_MAP = {
    'am'   : 'antiMalware',
    'ips'  : 'intrusionPrevention',
    'fw'   : 'firewall',
    'wr'   : 'webReputation',
    'im'   : 'integrityMonitoring',
    'li'   : 'logInspection',
    'appc' : 'applicationControl',
    'dc'   : 'deviceControl',
    'ac'   : 'activityMonitoring',
    'sap'  : 'SAP',
}

# --- Argument Parser ---
parser = argparse.ArgumentParser(description='Fetch settings from Trend Micro S&WP')
parser.add_argument('-p', '--policy', type=str, help='Policy name to filter (runs all policies if omitted)', default=None)
parser.add_argument('-m', '--module', nargs='*', help='Modules to fetch (e.g. ips am fw). Runs all if omitted.', default=None)
parser.add_argument('-o', '--output', choices=['csv', 'html', 'both'], default='both', help='Output format: csv, html, or both (default)')
parser.add_argument('-excluded_dir', action='store_true', help='List all excluded directories for the policy', default=False)
parser.add_argument('-excluded_files', action='store_true', help='List all excluded files for the policy', default=False)
parser.add_argument('-excluded_file_extensions', action='store_true', help='List all excluded file extensions for the policy', default=False)
parser.add_argument('-excluded_process_images', action='store_true', help='List all excluded process image files for the policy', default=False)
args = parser.parse_args()

if args.module:
    selected_modules = [MODULE_MAP[m] for m in args.module if m in MODULE_MAP]
else:
    selected_modules = list(MODULE_MAP.values())

AM_CONFIG_FIELDS = [
    'name', 'description', 'scanType',
    'documentExploitProtectionEnabled', 'documentExploitProtection', 'documentExploitHeuristicLevel',
    'machineLearningEnabled', 'behaviorMonitoringEnabled', 'documentRecoveryEnabled',
    'intelliTrapEnabled', 'memoryScanEnabled', 'spywareEnabled', 'alertEnabled',
    'directoriesToScan', 'directoryListID', 'filesToScan',
    'fileExtensionListID', 'excludedDirectoryListID', 'excludedFileListID',
    'excludedFileExtensionListID', 'excludedProcessImageFileListID',
    'realTimeScan', 'scanCompressedEnabled', 'scanCompressedMaximumSize',
    'scanCompressedMaximumLevels', 'scanCompressedMaximumFiles',
    'microsoftOfficeEnabled', 'microsoftOfficeLayers',
    'networkDirectoriesEnabled', 'customRemediationActionsEnabled', 'customScanActionsEnabled',
    'scanActionForVirus', 'scanActionForTrojans', 'scanActionForPacker',
    'scanActionForSpyware', 'scanActionForOtherThreats', 'scanActionForCookies',
    'scanActionForCVE', 'scanActionForHeuristics', 'scanActionForPossibleMalware',
    'amsiScanEnabled', 'scanActionForBehaviorMonitoring', 'scanActionForMachineLearning',
    'scanActionForAmsi', 'processMemoryScanAction',
    'detectionLevel', 'preventionLevel',
    'machineLearningDetectionLevel', 'machineLearningPreventionLevel',
    'behaviorMonitoringDetectionLevel', 'behaviorMonitoringPreventionLevel',
    'amsiDetectionLevel', 'amsiPreventionLevel',
    'processMemoryScanDetectionLevel', 'processMemoryScanPreventionLevel',
    'ID', 'cpuUsage'
]

LIST_ID_RESOLVERS = {
    'directoryListID':               ('directorylists',      'directories'),
    'excludedDirectoryListID':       ('directorylists',      'directories'),
    'excludedFileListID':            ('filelists',           'files'),
    'excludedFileExtensionListID':   ('fileextensionlists',  'fileExtensions'),
    'excludedProcessImageFileListID':('filelists',           'files'),
}

# --- Best Practices ---
_BP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'best_practices.json')
with open(_BP_FILE) as _f:
    BEST_PRACTICES = json.load(_f)['checks']

SEVERITY_ORDER = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}

def _val_matches(actual, recommended):
    """Case-insensitive string equality; also handles bool/int coercion."""
    return str(actual).strip().lower() == str(recommended).strip().lower()

def score_policy(policy_name, rows):
    """
    rows: list of [policy, module, subsection, setting, value]
    Returns: {
        'overall': {'score': float, 'passed': int, 'total': int},
        'by_module': {module: {'score': float, 'passed': int, 'total': int}},
        'findings': [{'module','subsection','setting','recommended','actual','severity','description','passed'}]
    }
    """
    # Build lookup: (module, subsection, setting) → value
    lookup = {}
    for _, module, subsection, setting, value in rows:
        lookup[(module, subsection, setting)] = value

    findings = []
    for check in [c for c in BEST_PRACTICES if c['module'] in selected_modules]:
        mod  = check['module']
        sub  = check['subsection']
        sett = check['setting']
        rec  = check['recommended']
        actual = lookup.get((mod, sub, sett), 'N/A')
        passed = _val_matches(actual, rec)
        findings.append({
            'module':      mod,
            'subsection':  sub,
            'setting':     sett,
            'recommended': rec,
            'actual':      actual,
            'severity':    check['severity'],
            'description': check['description'],
            'passed':      passed,
        })

    by_module = {}
    for f in findings:
        m = f['module']
        if m not in by_module:
            by_module[m] = {'passed': 0, 'total': 0}
        by_module[m]['total'] += 1
        if f['passed']:
            by_module[m]['passed'] += 1
    for m, v in by_module.items():
        v['score'] = round(100 * v['passed'] / v['total'], 1) if v['total'] else 0

    total  = len(findings)
    passed = sum(1 for f in findings if f['passed'])
    overall_score = round(100 * passed / total, 1) if total else 0

    return {
        'overall':   {'score': overall_score, 'passed': passed, 'total': total},
        'by_module': by_module,
        'findings':  findings,
    }

def grade(score):
    if score >= 90: return 'A'
    if score >= 80: return 'B'
    if score >= 70: return 'C'
    if score >= 60: return 'D'
    return 'F'

def score_color(score):
    if score >= 90: return '#155724'
    if score >= 70: return '#856404'
    return '#721c24'

def score_bg(score):
    if score >= 90: return '#d4edda'
    if score >= 70: return '#fff3cd'
    return '#f8d7da'

# --- Output accumulators ---
html_parts    = []
settings_rows = []   # rows for policy_settings_all.csv: [policy, module, subsection, setting, value]
scoring_results = {}  # policy_name → score_policy() result

def h(content):
    html_parts.append(content)

def c(policy, module, subsection, setting, value):
    settings_rows.append([policy, module, subsection, setting, value])

def html_table_rows(rows):
    out = ''
    for i, (label, value) in enumerate(rows):
        bg = ' style="background:#f9f9f9"' if i % 2 == 0 else ''
        out += f'<tr{bg}><td class="td-label">{label}</td><td class="td-value">{value}</td></tr>'
    return out

def html_settings_table(settings_dict, policy_settings):
    rows = [(label, policy_settings.get(key, {}).get('value', 'N/A')) for label, key in settings_dict.items()]
    return f'<table class="settings-table">{html_table_rows(rows)}</table>'

HTML_HEAD = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>S&WP Policy Report</title>
<style>
  body { font-family: Arial, sans-serif; font-size: 13px; color: #222; background: #f4f4f4; margin: 0; padding: 20px; }
  h1 { font-size: 20px; font-weight: 500; margin-bottom: 24px; }
  .policy-block { background: #fff; border: 1px solid #ddd; border-radius: 8px; margin-bottom: 24px; padding: 16px 20px; }
  .policy-title { font-size: 16px; font-weight: 500; margin: 0 0 16px; border-bottom: 1px solid #eee; padding-bottom: 8px; }
  details { border: 1px solid #e0e0e0; border-radius: 6px; margin-bottom: 8px; overflow: hidden; }
  summary { cursor: pointer; padding: 8px 12px; background: #f0f4f8; font-weight: 500; font-size: 13px; display: flex; justify-content: space-between; align-items: center; list-style: none; }
  summary::-webkit-details-marker { display: none; }
  summary:hover { background: #e4ecf4; }
  summary .badge { font-size: 11px; font-weight: 400; padding: 2px 8px; border-radius: 4px; }
  summary .badge.active { background: #d4edda; color: #155724; }
  summary .badge.inactive { background: #f8d7da; color: #721c24; }
  summary .badge.na { background: #e2e3e5; color: #383d41; }
  summary .state-label { font-size: 11px; color: #666; margin-right: 8px; }
  .module-content { padding: 12px; }
  .settings-table { width: 100%; border-collapse: collapse; font-size: 12px; }
  .td-label { color: #555; padding: 4px 8px; width: 45%; }
  .td-value { padding: 4px 8px; font-weight: 500; }
  .summary-block { margin-top: 16px; border-top: 1px solid #eee; padding-top: 12px; }
  .summary-title { font-size: 13px; font-weight: 500; margin-bottom: 8px; color: #444; }
  .summary-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 8px; }
  .summary-card { border: 1px solid #e0e0e0; border-radius: 6px; padding: 8px 12px; background: #fafafa; }
  .summary-card .mod-name { font-size: 11px; color: #666; margin-bottom: 2px; }
  .summary-card .mod-status { font-size: 12px; font-weight: 500; }
  .status-active { color: #155724; }
  .status-inactive { color: #721c24; }
  .status-na { color: #666; }
  .sub-section { font-size: 11px; font-weight: 500; color: #666; text-transform: uppercase; letter-spacing: 0.5px; margin: 10px 0 4px; }
  .score-block { margin-top: 16px; border-top: 1px solid #eee; padding-top: 14px; }
  .score-header { display: flex; align-items: center; gap: 16px; margin-bottom: 12px; }
  .score-circle { width: 60px; height: 60px; border-radius: 50%; display: flex; flex-direction: column; align-items: center; justify-content: center; font-weight: 700; }
  .score-circle .grade { font-size: 22px; line-height: 1; }
  .score-circle .pct   { font-size: 10px; line-height: 1; margin-top: 2px; }
  .score-meta { font-size: 12px; color: #555; }
  .score-meta strong { font-size: 14px; color: #222; display: block; margin-bottom: 2px; }
  .module-scores { display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 6px; margin-bottom: 12px; }
  .mod-score-card { border: 1px solid #e0e0e0; border-radius: 5px; padding: 6px 10px; background: #fafafa; }
  .mod-score-card .ms-name { font-size: 10px; color: #666; margin-bottom: 2px; }
  .mod-score-card .ms-bar-wrap { background: #eee; border-radius: 4px; height: 6px; margin: 4px 0; overflow: hidden; }
  .mod-score-card .ms-bar { height: 6px; border-radius: 4px; }
  .mod-score-card .ms-pct { font-size: 11px; font-weight: 600; }
  .findings-table { width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 8px; }
  .findings-table th { background: #f0f4f8; padding: 5px 8px; text-align: left; font-weight: 600; color: #444; }
  .findings-table td { padding: 4px 8px; border-bottom: 1px solid #f0f0f0; }
  .findings-table tr.fail-row { background: #fff8f8; }
  .sev-badge { display: inline-block; font-size: 10px; padding: 1px 6px; border-radius: 3px; font-weight: 600; }
  .sev-critical { background: #f8d7da; color: #721c24; }
  .sev-high     { background: #fde8c8; color: #7d4608; }
  .sev-medium   { background: #fff3cd; color: #856404; }
  .sev-low      { background: #e2e3e5; color: #383d41; }
  .chk-pass { color: #155724; font-weight: 700; }
  .chk-fail { color: #721c24; font-weight: 700; }
</style>
</head>
<body>
<h1>Trend Micro S&WP — Policy Report</h1>
'''

HTML_FOOT = '</body></html>'

# --- Helpers ---
def resolve_list_id(field, list_id):
    if not list_id or int(list_id) == 0:
        return 'None'
    endpoint, items_key = LIST_ID_RESOLVERS[field]
    r = requests.get(f'{BASE_URL}/{endpoint}/{list_id}', headers=headers)
    if r.ok:
        data  = r.json()
        name  = data.get('name', 'N/A')
        items = data.get(items_key) or data.get('items', [])
        print(f'    {field} → {name}')
        for item in items:
            print(f'      - {item}')
        return name
    print(f'    Error resolving {field} ID {list_id}: {r.status_code}')
    return f'Error({r.status_code})'

def fetch_am_config(config_id, config_type, policy_name):
    if not config_id or int(config_id) == 0:
        print(f'  {config_type}: No config assigned (ID=0)')
        return [], ''

    r = requests.get(f'{BASE_URL}/antimalwareconfigurations/{config_id}', headers=headers)
    if not r.ok:
        print(f'  Error fetching {config_type} ID {config_id}: {r.status_code} {r.text}')
        return [], ''

    config      = r.json()
    config_name = config.get('name', 'N/A')
    extracted   = {field: config.get(field, 'N/A') for field in AM_CONFIG_FIELDS}
    excl_html   = ''

    def fetch_list(endpoint, list_id, items_key, title):
        nonlocal excl_html
        if not list_id or int(list_id) == 0:
            print(f'  No {title} assigned')
            excl_html += f'<p style="color:#888;font-size:12px">No {title} assigned</p>'
            return
        r2 = requests.get(f'{BASE_URL}/{endpoint}/{list_id}', headers=headers)
        if r2.ok:
            data  = r2.json()
            lname = data.get('name', 'N/A')
            items = data.get(items_key) or data.get('items', [])
            print(f'  List Name: {lname}  Count: {len(items)}')
            for item in items:
                print(f'    - {item}')
                c(policy_name, 'antiMalware', config_type, title, item)
            rows_html = ''.join(f'<tr><td style="padding:2px 8px;font-size:12px">{item}</td></tr>' for item in items)
            excl_html += (
                f'<div class="sub-section">{title} — {lname} ({len(items)} items)</div>'
                f'<table class="settings-table">{rows_html}</table>'
            )
        else:
            print(f'  Error {r2.status_code}')
            excl_html += f'<p style="color:#c00;font-size:12px">Error fetching {title}: {r2.status_code}</p>'

    if args.excluded_dir:
        print(f'\n  --- Excluded Directories for [{config_type}] ---')
        fetch_list('directorylists', config.get('excludedDirectoryListID', 0), 'directories', 'Excluded Directories')

    if args.excluded_files:
        print(f'\n  --- Excluded Files for [{config_type}] ---')
        fetch_list('filelists', config.get('excludedFileListID', 0), 'files', 'Excluded Files')

    if args.excluded_file_extensions:
        print(f'\n  --- Excluded File Extensions for [{config_type}] ---')
        fetch_list('fileextensionlists', config.get('excludedFileExtensionListID', 0), 'fileExtensions', 'Excluded File Extensions')

    if args.excluded_process_images:
        print(f'\n  --- Excluded Process Image Files for [{config_type}] ---')
        fetch_list('filelists', config.get('excludedProcessImageFileListID', 0), 'files', 'Excluded Process Images')

    for field in LIST_ID_RESOLVERS:
        extracted[field] = resolve_list_id(field, config.get(field, 0))

    print(f'\n  [{config_type}] Policy: {policy_name} | Config: {config_name} (ID: {config_id})')
    for field, value in extracted.items():
        print(f'    {field}: {value}')

    return list(extracted.items()), excl_html


# --- Main ---
if __name__ == '__main__':
    r = requests.get(f'{BASE_URL}/policies', headers=headers)

    if r.ok:
        policies       = r.json()
        module_summary = []

        if args.policy:
            policies['policies'] = [p for p in policies.get('policies', []) if args.policy.lower() in p.get('name', '').lower()]
            if not policies['policies']:
                print(f'No policies found matching "{args.policy}"')
                exit(1)
            print(f'Running on policy: {args.policy}')
        else:
            print(f'No policy specified. Running on all policies. Total policies: {len(policies.get("policies", []))}')

        for p in policies.get('policies', []):
            name            = p.get('name', 'N/A')
            am              = p.get('antiMalware', {})
            ip              = p.get('intrusionPrevention', {})
            appc            = p.get('applicationControl', {})
            wr              = p.get('webReputation', {})
            dc              = p.get('deviceControl', {})
            ac              = p.get('activityMonitoring', {})
            fw              = p.get('firewall', {})
            li              = p.get('logInspection', {})
            im              = p.get('integrityMonitoring', {})
            policy_settings = p.get('policySettings', {})
            rec_scan_mode   = p.get('recommendationScanMode', 'N/A')

            module_summary.append({
                'policy': name,
                **{mod: p.get(mod, {}).get('moduleStatus', {}).get('status', 'N/A')
                   for mod in selected_modules}
            })

            # --- Terminal: Policy Header ---
            print(f'\n{"="*60}')
            print(f'Policy: {name}')
            print(f'{"="*60}')

            # --- HTML: Policy block open ---
            h(f'<div class="policy-block">')
            h(f'<div class="policy-title">{name}</div>')

            # --- Module Loop ---
            for module in selected_modules:
                data   = p.get(module, {})
                state  = data.get('state', 'N/A')
                status = data.get('moduleStatus', {}).get('status', 'N/A')

                print(f'\n  [{module}]')
                print(f'    state        : {state}')
                print(f'    moduleStatus : {status}')

                c(name, module, module, 'state', state)
                c(name, module, module, 'moduleStatus', status)

                badge_class = 'active' if status == 'active' else ('inactive' if status == 'inactive' else 'na')
                h(f'<details>')
                h(f'<summary><span>{module}</span><span><span class="state-label">state: {state}</span><span class="badge {badge_class}">{status}</span></span></summary>')
                h(f'<div class="module-content">')
                h(f'<table class="settings-table">')
                h(f'<tr><td class="td-label">state</td><td class="td-value">{state}</td></tr>')
                h(f'<tr style="background:#f9f9f9"><td class="td-label">moduleStatus</td><td class="td-value">{status}</td></tr>')

                # --- Anti-Malware ---
                if module == 'antiMalware':
                    sched_id   = am.get('realTimeScanScheduleID', 0)
                    sched_name = 'N/A'
                    if sched_id and int(sched_id) != 0:
                        sched_r = requests.get(f'{BASE_URL}/schedules/{sched_id}', headers=headers)
                        if sched_r.ok:
                            sched_name = sched_r.json().get('name', 'N/A')
                            print(f'    realTimeScanSchedule : {sched_name}')
                        else:
                            print(f'    realTimeScanSchedule : Error {sched_r.status_code}')
                    c(name, module, module, 'realTimeScanSchedule', sched_name)
                    h(f'<tr><td class="td-label">realTimeScanSchedule</td><td class="td-value">{sched_name}</td></tr>')
                    h('</table>')

                    for scan_type, config_key in [
                        ('realTimeScan',  'realTimeScanConfigurationID'),
                        ('manualScan',    'manualScanConfigurationID'),
                        ('scheduledScan', 'scheduledScanConfigurationID'),
                    ]:
                        rows, excl_html = fetch_am_config(am.get(config_key), scan_type, name)
                        if rows:
                            for label, value in rows:
                                c(name, module, scan_type, label, value)
                            h(f'<div class="sub-section">{scan_type}</div>')
                            h('<table class="settings-table">')
                            h(html_table_rows(rows))
                            h('</table>')
                            if excl_html:
                                h(excl_html)

                # --- Intrusion Prevention ---
                elif module == 'intrusionPrevention':
                    ip_rule_ids     = ip.get('ruleIDs', [])
                    inbound_tls     = policy_settings.get('intrusionPreventionSettingInspectInboundTlsTrafficEnabled',   {}).get('value', 'N/A')
                    outbound_tls    = policy_settings.get('intrusionPreventionSettingInspectOutboundTlsTrafficEnabled',  {}).get('value', 'N/A')
                    rec_mode        = policy_settings.get('intrusionPreventionSettingRecommendationScanMode',            {}).get('value', 'N/A')
                    auto_apply      = policy_settings.get('intrusionPreventionSettingAutoApplyRecommendationsEnabled',   {}).get('value', 'N/A')
                    auto_core_rules = policy_settings.get('intrusionPreventionSettingAutomaticallyApplyCoreIpsRules',    {}).get('value', 'N/A')
                    ips_first_hit   = policy_settings.get('intrusionPreventionSettingLogDataRuleFirstMatchEnabled',      {}).get('value', 'N/A')
                    auto_assign     = policy_settings.get('platformSettingAutoAssignNewIntrusionPreventionRulesEnabled', {}).get('value', 'N/A')

                    print(f'    ruleCount                  : {len(ip_rule_ids)}')
                    print(f'    TLS Inbound Inspection     : {inbound_tls}')
                    print(f'    TLS Outbound Inspection    : {outbound_tls}')
                    print(f'    Recommendation Scan Mode   : {rec_scan_mode}')
                    print(f'    IPS Rec Scan Mode Setting  : {rec_mode}')
                    print(f'    Auto-Apply Recommendations : {auto_apply}')
                    print(f'    Auto-Apply Core Rules      : {auto_core_rules}')
                    print(f'    Log First Match Only       : {ips_first_hit}')
                    print(f'    Auto-Assign New Rules      : {auto_assign}')

                    ips_rows = [
                        ('ruleCount',                  len(ip_rule_ids)),
                        ('TLS Inbound Inspection',     inbound_tls),
                        ('TLS Outbound Inspection',    outbound_tls),
                        ('Recommendation Scan Mode',   rec_scan_mode),
                        ('IPS Rec Scan Mode Setting',  rec_mode),
                        ('Auto-Apply Recommendations', auto_apply),
                        ('Auto-Apply Core Rules',      auto_core_rules),
                        ('Log First Match Only',       ips_first_hit),
                        ('Auto-Assign New Rules',      auto_assign),
                    ]
                    for label, value in ips_rows:
                        c(name, module, module, label, value)
                    h(html_table_rows(ips_rows))
                    h('</table>')

                # --- Application Control ---
                elif module == 'applicationControl':
                    trustRulesetID = appc.get('trustRulesetID', 0)
                    appc_rule_ids  = appc.get('ruleIDs', [])
                    enforce_mode   = appc.get('blockUnrecognized', 'N/A')
                    ruleset_name   = 'N/A'

                    if trustRulesetID and int(trustRulesetID) != 0:
                        r2 = requests.get(f'{BASE_URL}/applicationcontroltrustrulesets/{trustRulesetID}', headers=headers)
                        if r2.ok:
                            ruleset_name = r2.json().get('name', 'N/A')
                            print(f'    trustRuleset : {ruleset_name} (ID: {trustRulesetID})')
                        else:
                            print(f'    trustRuleset : Error {r2.status_code}')

                    print(f'    ruleCount        : {len(appc_rule_ids)}')
                    print(f'    Enforcement Mode : {enforce_mode}')

                    appc_rows = [
                        ('trustRuleset',     f'{ruleset_name} (ID: {trustRulesetID})'),
                        ('ruleCount',        len(appc_rule_ids)),
                        ('Enforcement Mode', enforce_mode),
                    ]
                    for label, value in appc_rows:
                        c(name, module, module, label, value)
                    h(html_table_rows(appc_rows))
                    h('</table>')

                # --- Web Reputation ---
                elif module == 'webReputation':
                    wr_settings = {
                        'Security Level'                      : 'webReputationSettingSecurityLevel',
                        'Block Untested Pages'                : 'webReputationSettingSecurityBlockUntestedPagesEnabled',
                        'Alerting Enabled'                    : 'webReputationSettingAlertingEnabled',
                        'Smart Protection Local Server'       : 'webReputationSettingSmartProtectionLocalServerEnabled',
                        'Smart Protection Local Server URLs'  : 'webReputationSettingSmartProtectionLocalServerUrls',
                        'Local Server Allow Off-Domain Global': 'webReputationSettingSmartProtectionLocalServerAllowOffDomainGlobal',
                        'Global Server Use Proxy'             : 'webReputationSettingSmartProtectionGlobalServerUseProxyEnabled',
                        'Global Server Proxy ID'              : 'webReputationSettingSmartProtectionWebReputationGlobalServerProxyId',
                        'Connection Lost Warning'             : 'webReputationSettingSmartProtectionServerConnectionLostWarningEnabled',
                        'Combined Mode Protection Source'     : 'webReputationSettingCombinedModeProtectionSource',
                        'Monitor Port List ID'                : 'webReputationSettingMonitorPortListId',
                        'Syslog Config ID'                    : 'webReputationSettingSyslogConfigId',
                        'Blocking Page Link'                  : 'webReputationSettingBlockingPageLink',
                        'Allowed URLs'                        : 'webReputationSettingAllowedUrls',
                        'Allowed URL Domains'                 : 'webReputationSettingAllowedUrlDomains',
                        'Blocked URLs'                        : 'webReputationSettingBlockedUrls',
                        'Blocked URL Domains'                 : 'webReputationSettingBlockedUrlDomains',
                        'Blocked URL Keywords'                : 'webReputationSettingBlockedUrlKeywords',
                    }
                    for label, key in wr_settings.items():
                        value = policy_settings.get(key, {}).get('value', 'N/A')
                        print(f'    {label:<40} : {value}')
                        c(name, module, module, label, value)
                    h('</table>')
                    h(html_settings_table(wr_settings, policy_settings))

                # --- Firewall ---
                elif module == 'firewall':
                    fw_rule_count      = len(fw.get('ruleIDs', []))
                    stateful_config_id = fw.get('globalStatefulConfigurationID', 0)
                    stateful_name      = 'N/A'
                    if stateful_config_id and int(stateful_config_id) != 0:
                        r2 = requests.get(f'{BASE_URL}/statefulconfigurations/{stateful_config_id}', headers=headers)
                        if r2.ok:
                            stateful_name = r2.json().get('name', 'N/A')
                            print(f'    Stateful Config : {stateful_name} (ID: {stateful_config_id})')
                        else:
                            print(f'    Stateful Config : Error {r2.status_code}')
                    print(f'    Firewall Rule Count : {fw_rule_count}')

                    c(name, module, module, 'Stateful Config', stateful_name)
                    c(name, module, module, 'Rule Count', fw_rule_count)
                    h(f'<tr><td class="td-label">Stateful Config</td><td class="td-value">{stateful_name}</td></tr>')
                    h(f'<tr style="background:#f9f9f9"><td class="td-label">Rule Count</td><td class="td-value">{fw_rule_count}</td></tr>')
                    h('</table>')

                    fw_settings = {
                        'Engine Options Enabled'           : 'firewallSettingEngineOptionsEnabled',
                        'Network Engine Mode'              : 'firewallSettingNetworkEngineMode',
                        'Network Engine Status Check'      : 'firewallSettingNetworkEngineStatusCheck',
                        'Combined Mode Protection Source'  : 'firewallSettingCombinedModeProtectionSource',
                        'Virtual & Container Network Scan' : 'firewallSettingVirtualAndContainerNetworkScanEnabled',
                        'Reconnaissance Enabled'           : 'firewallSettingReconnaissanceEnabled',
                        'Anti-Evasion Security Posture'    : 'firewallSettingAntiEvasionSecurityPosture',
                        'Failure Response Engine System'   : 'firewallSettingFailureResponseEngineSystem',
                        'Failure Response Packet Sanity'   : 'firewallSettingFailureResponsePacketSanityCheck',
                        'Syslog Config ID'                 : 'firewallSettingSyslogConfigId',
                    }
                    for label, key in fw_settings.items():
                        value = policy_settings.get(key, {}).get('value', 'N/A')
                        print(f'    {label:<40} : {value}')
                        c(name, module, module, label, value)
                    h(html_settings_table(fw_settings, policy_settings))

                # --- Device Control ---
                elif module == 'deviceControl':
                    dc_settings = {
                        'Control Mode'           : 'deviceControlSettingDeviceControlEnabled',
                        'Removable Storage Mode' : 'deviceControlSettingDeviceControlUsbStorageDeviceAction',
                        'Mobile Device Mode'     : 'deviceControlSettingMobileDeviceControlMode',
                        'Auto-Run Control Mode'  : 'deviceControlSettingDeviceControlAutoRunUsbAction',
                    }
                    for label, key in dc_settings.items():
                        value = policy_settings.get(key, {}).get('value', 'N/A')
                        print(f'    {label:<40} : {value}')
                        c(name, module, module, label, value)
                    h('</table>')
                    h(html_settings_table(dc_settings, policy_settings))

                # --- Activity Monitoring ---
                elif module == 'activityMonitoring':
                    detection_mode = policy_settings.get('activityMonitoringSettingDetectionMode', {}).get('value', 'N/A')
                    print(f'    Activity Monitoring Detection Mode : {detection_mode}')
                    c(name, module, module, 'Detection Mode', detection_mode)
                    h(f'<tr><td class="td-label">Detection Mode</td><td class="td-value">{detection_mode}</td></tr>')
                    h('</table>')

                # --- Log Inspection ---
                elif module == 'logInspection':
                    rule_ids  = li.get('ruleIDs', [])
                    auto_rec  = policy_settings.get('logInspectionSettingAutoApplyRecommendationsEnabled', {}).get('value', 'N/A')
                    syslog_id = policy_settings.get('logInspectionSettingSyslogConfigId', {}).get('value', 'N/A')
                    print(f'    Rule Count                 : {len(rule_ids)}')
                    print(f'    Auto-Apply Recommendations : {auto_rec}')
                    print(f'    Syslog Config ID           : {syslog_id}')
                    li_rows = [('Rule Count', len(rule_ids)), ('Auto-Apply Recommendations', auto_rec), ('Syslog Config ID', syslog_id)]
                    for label, value in li_rows:
                        c(name, module, module, label, value)
                    h(html_table_rows(li_rows))
                    h('</table>')

                # --- Integrity Monitoring ---
                elif module == 'integrityMonitoring':
                    rule_ids  = im.get('ruleIDs', [])
                    realtime  = policy_settings.get('integrityMonitoringSettingRealtimeEnabled', {}).get('value', 'N/A')
                    syslog_id = policy_settings.get('integrityMonitoringSettingSyslogConfigId', {}).get('value', 'N/A')
                    print(f'    Rule Count        : {len(rule_ids)}')
                    print(f'    Real-time Enabled : {realtime}')
                    print(f'    Syslog Config ID  : {syslog_id}')
                    im_rows = [('Rule Count', len(rule_ids)), ('Real-time Enabled', realtime), ('Syslog Config ID', syslog_id)]
                    for label, value in im_rows:
                        c(name, module, module, label, value)
                    h(html_table_rows(im_rows))
                    h('</table>')

                else:
                    h('</table>')

                h('</div></details>')

            # --- Terminal: Per-Policy Summary ---
            print(f'\n  {"─"*56}')
            print(f'  Summary: {name}')
            print(f'  {"─"*56}')
            for module in selected_modules:
                mod_state  = p.get(module, {}).get('state', 'N/A')
                mod_status = p.get(module, {}).get('moduleStatus', {}).get('status', 'N/A')
                print(f'    {module:<30} state={mod_state:<12} status={mod_status}')

            # --- HTML: Per-Policy Summary ---
            h('<div class="summary-block">')
            h('<div class="summary-title">Summary</div>')
            h('<div class="summary-grid">')
            for module in selected_modules:
                mod_state  = p.get(module, {}).get('state', 'N/A')
                mod_status = p.get(module, {}).get('moduleStatus', {}).get('status', 'N/A')
                status_cls = 'status-active' if mod_status == 'active' else ('status-inactive' if mod_status == 'inactive' else 'status-na')
                h(f'<div class="summary-card"><div class="mod-name">{module}</div><div class="mod-status {status_cls}">● {mod_status} &nbsp;<span style="font-weight:400;color:#666">({mod_state})</span></div></div>')
            h('</div></div>')

            # --- HTML: Best Practice Scoring ---
            policy_rows = [r for r in settings_rows if r[0] == name]
            result = score_policy(name, policy_rows)
            scoring_results[name] = result

            ov      = result['overall']
            bg      = score_bg(ov['score'])
            fg      = score_color(ov['score'])
            gr      = grade(ov['score'])

            h('<div class="score-block">')
            h('<div class="summary-title">Best Practice Compliance</div>')
            h('<div class="score-header">')
            h(f'<div class="score-circle" style="background:{bg};color:{fg}">'
              f'<span class="grade">{gr}</span>'
              f'<span class="pct">{ov["score"]}%</span>'
              f'</div>')
            h(f'<div class="score-meta"><strong>Overall Score: {ov["score"]}%</strong>'
              f'{ov["passed"]} of {ov["total"]} checks passed</div>')
            h('</div>')

            # per-module mini bars
            h('<div class="module-scores">')
            for mod, mv in sorted(result['by_module'].items()):
                bar_bg = score_bg(mv['score'])
                bar_fg = score_color(mv['score'])
                h(f'<div class="mod-score-card">'
                  f'<div class="ms-name">{mod}</div>'
                  f'<div class="ms-bar-wrap"><div class="ms-bar" style="width:{mv["score"]}%;background:{bar_fg}"></div></div>'
                  f'<div class="ms-pct" style="color:{bar_fg}">{mv["score"]}% ({mv["passed"]}/{mv["total"]})</div>'
                  f'</div>')
            h('</div>')

            # findings table (failed first, then sorted by severity)
            sorted_findings = sorted(result['findings'],
                                     key=lambda f: (f['passed'], SEVERITY_ORDER.get(f['severity'], 9)))
            h('<details><summary style="background:#f7f7f7;padding:7px 12px;cursor:pointer;font-size:12px;font-weight:500">▶ View All Checks</summary>')
            h('<table class="findings-table">')
            h('<tr><th>Module</th><th>Setting</th><th>Severity</th><th>Recommended</th><th>Actual</th><th>Status</th><th>Description</th></tr>')
            for f in sorted_findings:
                row_cls = '' if f['passed'] else ' class="fail-row"'
                status  = '<span class="chk-pass">✓ Pass</span>' if f['passed'] else '<span class="chk-fail">✗ Fail</span>'
                sev_cls = f'sev-{f["severity"]}'
                h(f'<tr{row_cls}>'
                  f'<td>{f["module"]}</td>'
                  f'<td>{f["setting"]}</td>'
                  f'<td><span class="sev-badge {sev_cls}">{f["severity"]}</span></td>'
                  f'<td>{f["recommended"]}</td>'
                  f'<td>{f["actual"]}</td>'
                  f'<td>{status}</td>'
                  f'<td style="color:#555">{f["description"]}</td>'
                  f'</tr>')
            h('</table></details>')
            h('</div>')

            h('</div>')

        # --- Timestamp suffix for all output files ---
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        # --- Module Status Summary (terminal only) ---
        if module_summary:
            all_modules = sorted({mod for row in module_summary for mod in row if mod != 'policy'})
            print(f'\nModule Status Summary:')
            print(','.join(['Policy'] + all_modules))
            for row in module_summary:
                print(','.join([row['policy']] + [row.get(m, 'N/A') for m in all_modules]))

        # --- Write combined XLSX ---
        if args.output in ('csv', 'both'):
            xlsx_file = f'swp_report_{ts}.xlsx'
            wb = openpyxl.Workbook()

            header_font  = Font(bold=True, color='FFFFFF')
            header_fill  = PatternFill('solid', fgColor='2E4A6E')
            pass_fill    = PatternFill('solid', fgColor='D4EDDA')
            fail_fill    = PatternFill('solid', fgColor='F8D7DA')
            sev_fills = {
                'critical': PatternFill('solid', fgColor='F8D7DA'),
                'high':     PatternFill('solid', fgColor='FDE8C8'),
                'medium':   PatternFill('solid', fgColor='FFF3CD'),
                'low':      PatternFill('solid', fgColor='E2E3E5'),
            }

            def style_header(ws, headers):
                ws.append(headers)
                for cell in ws[1]:
                    cell.font       = header_font
                    cell.fill       = header_fill
                    cell.alignment  = Alignment(horizontal='center', vertical='center')

            def auto_width(ws):
                for col in ws.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            # Sheet 1 — Module Status Summary
            ws1 = wb.active
            ws1.title = 'Module Status'
            if module_summary:
                style_header(ws1, ['Policy'] + all_modules)
                for row in module_summary:
                    ws1.append([row['policy']] + [row.get(m, 'N/A') for m in all_modules])
            auto_width(ws1)

            # Sheet 2 — All Settings
            ws2 = wb.create_sheet('All Settings')
            style_header(ws2, ['Policy', 'Module', 'SubSection', 'Setting', 'Value'])
            for row in settings_rows:
                ws2.append(row)
            auto_width(ws2)

            # Sheet 3 — Recommendations
            ws3 = wb.create_sheet('Recommendations')
            style_header(ws3, ['Policy', 'Module', 'Setting', 'Severity', 'Recommended', 'Actual', 'Status', 'Description'])
            if scoring_results:
                for pol_name, result in scoring_results.items():
                    for f in sorted(result['findings'],
                                    key=lambda x: (x['passed'], SEVERITY_ORDER.get(x['severity'], 9))):
                        ws3.append([
                            pol_name, f['module'], f['setting'], f['severity'],
                            f['recommended'], f['actual'],
                            'PASS' if f['passed'] else 'FAIL',
                            f['description'],
                        ])
                        row_idx = ws3.max_row
                        status_cell = ws3.cell(row=row_idx, column=7)
                        sev_cell    = ws3.cell(row=row_idx, column=4)
                        status_cell.fill = pass_fill if f['passed'] else fail_fill
                        status_cell.font = Font(bold=True)
                        sev_cell.fill    = sev_fills.get(f['severity'], PatternFill())
            auto_width(ws3)

            wb.save(xlsx_file)
            print(f'\nReport saved to {xlsx_file} (sheets: Module Status, All Settings, Recommendations)')

        # --- Write HTML file ---
        if args.output in ('html', 'both'):
            # Build scoring dashboard to prepend before policy blocks
            dashboard_parts = []
            if scoring_results:
                dashboard_parts.append('<div class="policy-block">')
                dashboard_parts.append('<div class="policy-title">Best Practice Compliance Dashboard</div>')
                dashboard_parts.append('<table class="findings-table">')
                dashboard_parts.append('<tr><th>Policy</th><th>Score</th><th>Grade</th><th>Passed</th><th>Failed</th><th>Critical Fails</th></tr>')
                for pol_name, result in scoring_results.items():
                    ov       = result['overall']
                    g        = grade(ov['score'])
                    fg       = score_color(ov['score'])
                    bg       = score_bg(ov['score'])
                    failed   = ov['total'] - ov['passed']
                    crit_fail= sum(1 for f in result['findings'] if not f['passed'] and f['severity'] == 'critical')
                    dashboard_parts.append(
                        f'<tr>'
                        f'<td>{pol_name}</td>'
                        f'<td><strong style="color:{fg}">{ov["score"]}%</strong></td>'
                        f'<td><span style="background:{bg};color:{fg};padding:2px 8px;border-radius:4px;font-weight:700">{g}</span></td>'
                        f'<td style="color:#155724">{ov["passed"]}</td>'
                        f'<td style="color:#721c24">{failed}</td>'
                        f'<td style="color:#721c24;font-weight:{"700" if crit_fail else "400"}">{crit_fail}</td>'
                        f'</tr>'
                    )
                dashboard_parts.append('</table></div>')

            html_file = f'policy_report_{ts}.html'
            with open(html_file, 'w') as hf:
                hf.write(HTML_HEAD)
                hf.write('\n'.join(dashboard_parts))
                hf.write('\n'.join(html_parts))
                hf.write(HTML_FOOT)
            print(f'HTML report saved to {html_file}')

    else:
        print(f'Error {r.status_code}: {r.text}')
