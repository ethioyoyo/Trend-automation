import requests
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL = 'https://app.deepsecurity.trendmicro.com/api'
API_KEY  = '8BEC6851-96F0-D7D3-5002-C1062BC9E89C:CB77AF3B-16F2-F519-1E61-AFF9B2B2CF46:qKhC1TxhgCVVze6qISiL2oJq1f9GcEz+KJH2iBVUVBE='

headers = {
    'api-secret-key': API_KEY,
    'api-version': 'v1',
    'Content-Type': 'application/json'
}

# --- Argument Parser ---
parser = argparse.ArgumentParser(description='Fetch AntiMalware settings from Trend Micro S&WP')
parser.add_argument('-p',                        type=str,            default=None,  help='Policy name to filter (runs all if omitted)')
parser.add_argument('-excluded_dir',             action='store_true', default=False, help='List all excluded directories')
parser.add_argument('-excluded_files',           action='store_true', default=False, help='List all excluded files')
parser.add_argument('-excluded_file_extensions', action='store_true', default=False, help='List all excluded file extensions')
parser.add_argument('-excluded_process_images',  action='store_true', default=False, help='List all excluded process image files')
args = parser.parse_args()

AM_CONFIG_FIELDS = [
    'name', 'description', 'scanType',
    'documentExploitProtectionEnabled', 'documentExploitProtection', 'documentExploitHeuristicLevel',
    'machineLearningEnabled', 'behaviorMonitoringEnabled', 'documentRecoveryEnabled',
    'intelliTrapEnabled', 'memoryScanEnabled', 'spywareEnabled', 'alertEnabled',
    'directoriesToScan', 'directoryListID', 'filesToScan',
    'fileExtensionListID', 'excludedDirectoryListID', 'excludedFileListID',
    'excludedFileExtensionListID', 'excludedProcessImageFileListID',
    'realTimeScan', 'scanCompressedEnabled', 'scanCompressedMaximumSize',
    'scanCompressedMaximumLevels', 'scanCompressedMaximumFiles',
    'microsoftOfficeEnabled', 'microsoftOfficeLayers',
    'networkDirectoriesEnabled', 'customRemediationActionsEnabled', 'customScanActionsEnabled',
    'scanActionForVirus', 'scanActionForTrojans', 'scanActionForPacker',
    'scanActionForSpyware', 'scanActionForOtherThreats', 'scanActionForCookies',
    'scanActionForCVE', 'scanActionForHeuristics', 'scanActionForPossibleMalware',
    'amsiScanEnabled', 'scanActionForBehaviorMonitoring', 'scanActionForMachineLearning',
    'scanActionForAmsi', 'processMemoryScanAction',
    'detectionLevel', 'preventionLevel',
    'machineLearningDetectionLevel', 'machineLearningPreventionLevel',
    'behaviorMonitoringDetectionLevel', 'behaviorMonitoringPreventionLevel',
    'amsiDetectionLevel', 'amsiPreventionLevel',
    'processMemoryScanDetectionLevel', 'processMemoryScanPreventionLevel',
    'ID', 'cpuUsage'
]

LIST_ID_RESOLVERS = {
    'directoryListID':                ('directorylists',     'items'),
    'excludedDirectoryListID':        ('directorylists',     'items'),
    'excludedFileListID':             ('filelists',          'items'),
    'excludedFileExtensionListID':    ('fileextensionlists', 'items'),
    'excludedProcessImageFileListID': ('filelists',          'items'),
}

# --- Styling helpers ---
HEADER_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HEADER_FILL    = PatternFill('solid', start_color='1F4E79')
ALT_ROW_FILL   = PatternFill('solid', start_color='D9E1F2')
CENTER_ALIGN   = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN

def style_data_rows(ws, start_row=2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=0):
        # Only style rows that have at least one value
        if not any(cell.value for cell in row):
            continue
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        for cell in row:
            cell.alignment = LEFT_ALIGN
            if fill:
                cell.fill = fill

def auto_width(ws, max_width=50):
    for col in ws.columns:
        width = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, max_width)

# --- List resolver ---
def resolve_list(endpoint, list_id, items_key):
    """Fetch list by ID, return (name, items). Returns ('None', []) if ID is 0."""
    if not list_id or int(list_id) == 0:
        return 'None', []
    r = requests.get(f'{BASE_URL}/{endpoint}/{list_id}', headers=headers)
    if r.ok:
        data = r.json()
        name  = data.get('name', 'N/A')
        items = data.get(items_key, [])
        return name, items  
    return f'Error({r.status_code})', []

# --- Main AM config fetcher ---
def fetch_am_config(config_id, config_type, policy_name, ws_settings, ws_excl):
    if not config_id or int(config_id) == 0:
        print(f'  {config_type}: No config assigned (ID=0)')
        return

    resp = requests.get(f'{BASE_URL}/antimalwareconfigurations/{config_id}', headers=headers)
    if not resp.ok:
        print(f'  Error fetching {config_type} ID {config_id}: {resp.status_code}')
        return

    config      = resp.json()
    config_name = config.get('name', 'N/A')
    extracted   = {field: config.get(field, 'N/A') for field in AM_CONFIG_FIELDS}

    # Resolve all list ID fields → names (single API call each, stored once)
    resolved_lists = {}
    for field, (endpoint, items_key) in LIST_ID_RESOLVERS.items():
        name, items = resolve_list(endpoint, config.get(field, 0), items_key)
        extracted[field]      = name
        resolved_lists[field] = (name, items)

    # --- Write to Settings sheet ---
    ws_settings.append(
        [policy_name, config_type, config_id, config_name] +
        [extracted.get(f, '') for f in AM_CONFIG_FIELDS]
    )

    # --- Write exclusion lists to Exclusions sheet (only if flags passed) ---
    flag_map = {
        'excludedDirectoryListID':        args.excluded_dir,
        'excludedFileListID':             args.excluded_files,
        'excludedFileExtensionListID':    args.excluded_file_extensions,
        'excludedProcessImageFileListID': args.excluded_process_images,
    }

    # Change this debug line
    print(f'  DEBUG resolved_lists: { {k: (n, len(i)) for k, (n, i) in resolved_lists.items()} }')

    # To this — shows actual items
    print(f'  DEBUG resolved_lists: { {k: (n, i) for k, (n, i) in resolved_lists.items()} }')
    for field, flag in flag_map.items():
        if flag:
            list_name, items = resolved_lists[field]
            print(f'  Writing {len(items)} items for {field} to sheet...') 
            print(f'\n  --- {field} [{config_type}] ---')
            print(f'  List Name : {list_name}')
            print(f'  Count     : {len(items)}')
            for item in items:
                print(f'    - {item}')
                ws_excl.append([policy_name, config_type, config_name, field, list_name, item])

    print(f'  [{config_type}] {policy_name} | Config: {config_name} (ID: {config_id})')


# ===================== MAIN =====================
r = requests.get(f'{BASE_URL}/policies', headers=headers)

if not r.ok:
    print(f'Error {r.status_code}: {r.text}')
    exit(1)

policies = r.json().get('policies', [])

if args.p:
    policies = [p for p in policies if args.p.lower() in p.get('name', '').lower()]
    if not policies:
        print(f'No policies found matching "{args.p}"')
        exit(1)
    print(f'Running on policy filter: "{args.p}" — {len(policies)} match(es)')
else:
    print(f'No filter — running on all {len(policies)} policies')

# --- Build Workbook ---
wb = Workbook()

# Sheet 1 — AM Settings
ws_settings = wb.active
ws_settings.title = 'AM Settings'
ws_settings.append(
    ['Policy Name', 'Config Type', 'Config ID', 'Config Name'] + AM_CONFIG_FIELDS
)

# Sheet 2 — Exclusion Lists
ws_excl = wb.create_sheet('Exclusion Lists')
ws_excl.append(['Policy Name', 'Config Type', 'Config Name', 'List Field', 'List Name', 'Item'])

# Sheet 3 — Summary
ws_summary = wb.create_sheet('Summary')
ws_summary.append(['Policy Name', 'State', 'Module Status', 'RealTime Config ID', 'Manual Config ID', 'Scheduled Config ID', 'Schedule ID'])

# --- Process policies ---
for p in policies:
    name = p.get('name', 'N/A')
    am   = p.get('antiMalware', {})

    module_status = am.get('moduleStatus', {}).get('status', 'N/A')
    ws_summary.append([
        name,
        am.get('state', 'N/A'),
        module_status,
        am.get('realTimeScanConfigurationID', 0),
        am.get('manualScanConfigurationID', 0),
        am.get('scheduledScanConfigurationID', 0),
        am.get('realTimeScanScheduleID', 0),
    ])

    print(f'\n=== Policy: {name} | State: {am.get("state")} | Module: {module_status} ===')

    fetch_am_config(am.get('realTimeScanConfigurationID'),  'realTimeScan',  name, ws_settings, ws_excl)
    fetch_am_config(am.get('manualScanConfigurationID'),    'manualScan',    name, ws_settings, ws_excl)
    fetch_am_config(am.get('scheduledScanConfigurationID'), 'scheduledScan', name, ws_settings, ws_excl)

    sched_id = am.get('realTimeScanScheduleID', 0)
    if sched_id and int(sched_id) != 0:
        sr = requests.get(f'{BASE_URL}/schedules/{sched_id}', headers=headers)
        if sr.ok:
            print(f'  realTimeScanScheduleID → {sr.json().get("name")}')

# --- Confirm row counts BEFORE styling ---
print(f'ws_settings rows: {ws_settings.max_row}')
print(f'ws_excl rows:     {ws_excl.max_row}')
print(f'ws_summary rows:  {ws_summary.max_row}')

# --- Apply styling AFTER all data written ---
for ws in [ws_settings, ws_excl, ws_summary]:
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    ws.freeze_panes = 'A2'

# --- Save ---
output_file = f'antimalware_{"_".join(args.p.split()) if args.p else "all"}.xlsx'
wb.save(output_file)
print(f'\nDone — saved to {output_file}')